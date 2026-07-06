"""Sanity tests for fill.py, run with plain CPython (pip install openpyxl)."""

import glob
import io
import openpyxl
import fill

EMPTY = sorted(f for f in glob.glob("cv_examples/*.xlsx") if "template" not in f)


def _sheet(data):
    return openpyxl.load_workbook(io.BytesIO(data)).active


def test_analyze():
    for f in EMPTY:
        info = fill.analyze(open(f, "rb").read())
        assert info["month"] and info["year"], info
        assert info["workdays"] * fill.DAILY_CAP == info["capacity"]
        print(f"  analyze {info['month']} {info['year']}: "
              f"{info['workdays']} workdays, holidays={len(info['holidays'])}")


def test_fill_totals_and_token():
    for f in EMPTY:
        src = open(f, "rb").read()
        before = _sheet(src)
        out = fill.fill(src, [{"wp": "WP3", "hours": 40}, {"wp": "WP4", "hours": 20}])
        ws = _sheet(out)
        total = sum(ws.cell(r, fill.COL_RESEARCH).value or 0 for r in range(13, 44))
        assert total == 60, total
        # per-day cap respected
        assert all((ws.cell(r, fill.COL_RESEARCH).value or 0) <= fill.DAILY_CAP for r in range(13, 44))
        # integrity token + a formula preserved
        assert ws["O2"].value == before["O2"].value, "O2 token changed!"
        assert ws["C44"].value == "=SUM(C13:C43)"
        # holidays untouched
        for h in fill.analyze(src)["holidays"]:
            row = 12 + h["day"]
            assert ws.cell(row, fill.COL_RESEARCH).value in (None, 0, ""), "wrote on a holiday!"
        print(f"  fill 60h into {before['B10'].value}: total ok, token ok")


def test_multi_wp_blocks():
    # Two WPs must occupy separate, non-overlapping day-blocks with their own labels.
    src = open(EMPTY[2], "rb").read()  # July: no holidays, 23 workdays
    out = fill.fill(src, [{"wp": "WP3", "hours": 24}, {"wp": "WP4", "hours": 16}])
    ws = _sheet(out)
    by_wp = {}
    for r in range(13, 44):
        note = ws.cell(r, fill.COL_NOTE).value
        hrs = ws.cell(r, fill.COL_RESEARCH).value
        if hrs:
            by_wp.setdefault(note, []).append((r, hrs))
    assert set(by_wp) == {"WP3", "WP4"}, by_wp
    assert sum(h for _, h in by_wp["WP3"]) == 24
    assert sum(h for _, h in by_wp["WP4"]) == 16
    wp3_rows = [r for r, _ in by_wp["WP3"]]
    wp4_rows = [r for r, _ in by_wp["WP4"]]
    assert set(wp3_rows).isdisjoint(wp4_rows), "WP day-blocks overlap"
    assert max(wp3_rows) < min(wp4_rows), "WP blocks not in order"
    print(f"  multi-WP: WP3 on {len(wp3_rows)} days, WP4 on {len(wp4_rows)} days, separate")


def test_overflow_guard():
    src = open(EMPTY[0], "rb").read()
    try:
        fill.fill(src, [{"wp": "WP3", "hours": 100000}])
    except ValueError as e:
        print(f"  overflow guarded: {str(e)[:50]}…")
    else:
        raise AssertionError("overflow not caught")


if __name__ == "__main__":
    test_analyze()
    test_fill_totals_and_token()
    test_multi_wp_blocks()
    test_overflow_guard()
    print("all tests passed")
