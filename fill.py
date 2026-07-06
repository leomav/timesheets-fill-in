"""Timesheet fill logic.

Runs identically in plain CPython (for tests) and in the browser via Pyodide.
Only cell *values* are written, so openpyxl preserves every other feature of the
workbook untouched -- styling, the K/M formulas, and critically the O2/O1
integrity token the institution stamps on each sheet.
"""

import io
import math
import openpyxl

# Layout constants for the NTUA / EUREKA3D-XR monthly timesheet.
FIRST_ROW = 13          # first day row
LAST_ROW = 43           # last possible day row (31-day month)
COL_DATE = 1            # A
COL_DAY = 2            # B  (Greek day name)
COL_RESEARCH = 3        # C  (hours land here)
COL_NOTE = 10           # J  (WP label / pre-marked holiday)
COL_MONTH_CELL = "B10"
COL_YEAR_CELL = "B9"
DAILY_CAP = 8

WEEKEND = {"Σαββάτο", "Κυριακή"}


def _workday_rows(ws):
    """Rows that are real working days: a weekday with an empty note column.

    A non-empty note in column J means the sheet already flags that day as a
    public holiday (e.g. 'Πρωτομαγιά', 'Αγίου Πνεύματος'), so it is skipped.
    """
    rows = []
    for r in range(FIRST_ROW, LAST_ROW + 1):
        if ws.cell(r, COL_DATE).value is None:
            continue
        if ws.cell(r, COL_DAY).value in WEEKEND:
            continue
        if ws.cell(r, COL_NOTE).value not in (None, ""):
            continue
        rows.append(r)
    return rows


def _even_spread(hours, ndays):
    """Split `hours` across `ndays` as evenly as possible, each day <= DAILY_CAP."""
    base = hours // ndays
    rem = hours - base * ndays
    return [base + (1 if i < rem else 0) for i in range(ndays)]


def _allocate_days(hours_list, total_days, cap):
    """Decide how many working days each work package gets.

    Days are handed out to fill *all* available working days when possible, so
    the hours land spread thin-and-even rather than piled onto the first days.
    Each package gets between ceil(h/cap) days (so no day exceeds the cap) and
    h days (so no day drops to zero hours). Extra days go to the packages with
    the largest fair-share remainder.
    """
    n = len(hours_list)
    total_hours = sum(hours_list)
    if total_hours == 0:
        return [0] * n

    exact = [h / total_hours * total_days for h in hours_list]
    days = [max(math.ceil(h / cap), 1) if h else 0 for h in hours_list]

    remaining = total_days - sum(days)
    # If we've already committed more days than exist, callers guard against the
    # infeasible case earlier; just return the minimum feasible plan.
    if remaining <= 0:
        return days

    order = sorted(range(n), key=lambda i: exact[i] - int(exact[i]), reverse=True)
    while remaining > 0:
        progressed = False
        for i in order:
            if remaining <= 0:
                break
            if days[i] < hours_list[i]:  # keep at least 1h per day
                days[i] += 1
                remaining -= 1
                progressed = True
        if not progressed:  # every package is already at 1h/day; stop
            break
    return days


def analyze(data: bytes) -> dict:
    """Read a workbook and report month/year and how many working days it has."""
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    rows = _workday_rows(ws)
    holidays = []
    for r in range(FIRST_ROW, LAST_ROW + 1):
        note = ws.cell(r, COL_NOTE).value
        if note not in (None, "") and ws.cell(r, COL_DAY).value not in WEEKEND:
            holidays.append({"day": ws.cell(r, COL_DATE).value, "note": note})
    return {
        "month": ws[COL_MONTH_CELL].value,
        "year": ws[COL_YEAR_CELL].value,
        "workdays": len(rows),
        "capacity": len(rows) * DAILY_CAP,
        "holidays": holidays,
    }


def fill(data: bytes, allocations: list) -> bytes:
    """Fill a workbook.

    `allocations` is a list of {"wp": str, "hours": number}. Each WP is placed
    on its own consecutive block of working days (label in column J, hours in
    column C), even-spread and capped at DAILY_CAP per day.
    Returns the new .xlsx as bytes. Raises ValueError if it cannot fit.
    """
    allocations = [a for a in allocations if a.get("hours")]
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    rows = _workday_rows(ws)

    labels, hours_list = [], []
    for a in allocations:
        hours = int(a["hours"])
        if hours <= 0:
            continue
        labels.append(str(a["wp"]).strip())
        hours_list.append(hours)

    total = sum(hours_list)
    if total > len(rows) * DAILY_CAP:
        raise ValueError(
            f"Needs {total}h but the month only has {len(rows)} working days "
            f"({len(rows) * DAILY_CAP}h capacity)."
        )

    day_counts = _allocate_days(hours_list, len(rows), DAILY_CAP)

    idx = 0
    for wp, hours, ndays in zip(labels, hours_list, day_counts):
        block = rows[idx:idx + ndays]
        idx += ndays
        for row, hrs in zip(block, _even_spread(hours, ndays)):
            ws.cell(row, COL_RESEARCH).value = hrs
            ws.cell(row, COL_NOTE).value = wp

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
